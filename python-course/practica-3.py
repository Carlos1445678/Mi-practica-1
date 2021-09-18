from pyhunter import PyHunter
from openpyxl import Workbook
import getpass
import openpyxl

def Busqueda(organizacion):

    #Cantidad de resultados esperados de la búsqueda
    #El límite MENSUAL de Hunter es 50, cuidado!
    print(organizacion)
    resultado = hunter.domain_search (organizacion,hunter)
    
    #resultado = hunter.domain_search (company = organizacion, limit= 1 ,emails_type ='personal')
    return resultado

def GuardarInformacion(organizacion):
 excel_document = openpyxl.load_workbook('Hunteryoutube.com.xlsx')
 print type(excel_document)
  #def GuardarInformacion(datosEncontrados,organizacion):
    #libro = Workbook()
    #hoja=libro.active
    #hoja["A1,A123"]=(organizacion,hunter)
    #libro.save("Hunteryoutube.com.xlsx")
    #print(organizacion)
    #print(datosEncontrados)
    #hoja = libro.create_sheet(organizacion)
    #libro.save(hunter + organizacion + ".xlsx")
    #Agrega el codigo necesario para guardar en formato tabla
    #dentro del libro de Excel, información que consideres relevante
    #de lo obtenido en la búsqueda.
    libro.save("Hunter" + organizacion + ".xlsx")
    

print("Script para buscar información")
apikey = getpass.getpass("Ingresa tu API key: ")
hunter = PyHunter (apikey)
orga = input("Dominio a investigar: ")
datosEncontrados = Busqueda(orga)
if datosEncontrados == None:
    exit()
else:
    print(datosEncontrados)
    print(type(datosEncontrados))
    GuardarInformacion(orga)
    

